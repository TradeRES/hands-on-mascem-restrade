# ----------- IMPORTS ----------- #

import os
import sys
import json
import glob
import subprocess
try:
  import pandas as pd
except:
  subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'pandas'])
  import pandas as pd
try:
  import natsort
except:
  subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'natsort'])
  import natsort
from openpyxl import load_workbook
from openpyxl.chart import (
  BarChart,
  LineChart,
  PieChart,
  Reference
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment

# ----------- FUNCTIONS ----------- #

# Read JSON file
def read_json(filepath):
  f = open(filepath, 'r')
  json_content = json.load(f)
  f.close()
  return json_content

# Write JSON file
def write_json(json_content, filepath):
  f = open(filepath, 'w')
  json.dump(json_content, f, indent = 2)
  f.close()

# Save to XLSX file
def save_df_to_xlsx_file(file_name, sheet_names, df_list):
  # If lens are equal
  if len(sheet_names) == len(df_list):
    #
    # Save data
    try:
      #
      # Safety check to set xlsx writer
      if not os.path.exists(file_name):
        writer = pd.ExcelWriter(file_name, engine = 'openpyxl')
      else:
        writer = pd.ExcelWriter(file_name, engine = 'openpyxl', mode = 'a', if_sheet_exists='overlay')
      # 
      # Set sheets
      n_sheets = len(sheet_names)
      for i in range(n_sheets):
        df_list[i].to_excel(writer, sheet_name = sheet_names[i], index = False)
      # 
      # Save
      writer.close()
    except Exception as ex:
      print('Error while saving data to xlsx:', ex)
  else:
    print('The number of sheet_names elements must match the number of df_list elements.', sheet_names, 'not saved.')

def draw_charts_xlsx_file(file_name):
  # If file exists
  if os.path.exists(file_name):
    try:
      #
      # Read xlsx
      wb = load_workbook(filename = file_name)
      #
      # Market Price sheet
      mp_ws = wb['Market Prices']
      # DEBUG: print(ws['B2':'E25'], ws['B1'].value)
      #
      # Draw market prices line chart
      mp_chart = LineChart()
      mp_chart.height = 10 # default is 7.5
      mp_chart.width = 20 # default is 15
      mp_chart.title = 'Session Market Prices '
      mp_chart.style = 11
      mp_chart.y_axis.title = 'EUR/MWh'
      mp_chart.y_axis.scaling.max = 120
      mp_chart.x_axis.title = 'Period'
      #
      mp_data = Reference(mp_ws, min_col = 2, min_row = 1, max_col = mp_ws.max_column, max_row = mp_ws.max_row)
      mp_chart.add_data(mp_data, titles_from_data = True)
      #
      pt_series = mp_chart.series[0]
      pt_series.graphicalProperties.line.solidFill = '145a32'
      pt_series.graphicalProperties.line.width = 33300 # width in EMUs
      es_series = mp_chart.series[1]
      es_series.graphicalProperties.line.solidFill = 'F39C12'
      es_series.graphicalProperties.line.width = 33300 # width in EMUs
      mi_series = mp_chart.series[2]
      mi_series.graphicalProperties.line.solidFill = '581845'
      mi_series.graphicalProperties.line.width = 33300 # width in EMUs
      #
      #mp_ws.add_chart(mp_chart, 'F2')
      #
      #
      # Mean Prices sheet
      ap_ws = wb['Mean Prices']
      # Draw Mean Price sheet
      ap_chart = BarChart()
      ap_chart.height = 10 # default is 7.5
      ap_chart.width = 20 # default is 15
      ap_chart.type = 'col'
      ap_chart.style = 11
      ap_chart.title = 'Mean Prices'
      ap_chart.y_axis.title = 'EUR/MWh'
      #
      # Merge cells & center
      ap_ws.merge_cells(range_string = 'A2:A4')
      cell = ap_ws.cell(row = 2, column = 1)
      cell.alignment = Alignment(horizontal='center', vertical='center')
      #
      ap_data = Reference(ap_ws, min_col = 3, min_row = 1, max_col = ap_ws.max_column, max_row = ap_ws.max_row)
      ap_chart.add_data(ap_data, titles_from_data = True)
      ap_labels = Reference(ap_ws, min_col = 1, min_row = 2, max_col = 2, max_row = ap_ws.max_row)
      ap_chart.set_categories(ap_labels)
      #
      pt_series = ap_chart.series[0]
      pt_series.graphicalProperties.solidFill = '145a32'
      pt_series.graphicalProperties.line.solidFill = '145a32'
      es_series = ap_chart.series[1]
      es_series.graphicalProperties.solidFill = 'F39C12'
      es_series.graphicalProperties.line.solidFill = 'F39C12'
      mi_series = ap_chart.series[2]
      mi_series.graphicalProperties.solidFill = '581845'
      mi_series.graphicalProperties.line.solidFill = '581845'
      #
      #ap_ws.add_chart(ap_chart, 'G2')
      #
      #
      # Traded Volume sheet
      tv_ws = wb['Traded Volume']
      # Draw energy volume bars chart
      tv_chart = BarChart()
      tv_chart.height = 10 # default is 7.5
      tv_chart.width = 20 # default is 15
      tv_chart.type = 'col'
      tv_chart.style = 11
      tv_chart.grouping = 'stacked'
      tv_chart.overlap = 100
      tv_chart.title = 'Traded Volume'
      tv_chart.y_axis.title = 'MWh'
      tv_chart.x_axis.title = 'Period'
      #
      tv_data = Reference(tv_ws, min_col = 2, min_row = 1, max_col = tv_ws.max_column, max_row = tv_ws.max_row)
      tv_chart.add_data(tv_data, titles_from_data = True)
      #
      pt_series = tv_chart.series[0]
      pt_series.graphicalProperties.solidFill = '145a32'
      pt_series.graphicalProperties.line.solidFill = '145a32'
      es_series = tv_chart.series[1]
      es_series.graphicalProperties.solidFill = 'F39C12'
      es_series.graphicalProperties.line.solidFill = 'F39C12'
      mi_series = tv_chart.series[2]
      mi_series.graphicalProperties.solidFill = '581845'
      mi_series.graphicalProperties.line.solidFill = '581845'
      #
      #tv_ws.add_chart(tv_chart, 'F2')
      #
      #
      # Total Traded Volume by Country
      ttvc_ws = wb['Total Traded Volume by Country']
      # Draw energy volume bars chart
      ttvc_chart = PieChart()
      ttvc_chart.height = 10 # default is 7.5
      ttvc_chart.width = 20 # default is 15
      tv_chart.style = 11
      ttvc_labels = Reference(ttvc_ws, min_col = 1, min_row = 2, max_row = 3)
      ttvc_data = Reference(ttvc_ws, min_col = 2, min_row = 1, max_col = 3, max_row = 3)
      ttvc_chart.add_data(ttvc_data, titles_from_data = True)
      ttvc_chart.set_categories(ttvc_labels)
      ttvc_chart.title = 'Total Traded Volume by Country'
      ttvc_chart.dataLabels = DataLabelList()
      ttvc_chart.dataLabels.showVal = True
      #
      pt_series = ttvc_chart.series[0]
      pt_point = DataPoint(idx = 0)
      pt_point.graphicalProperties.solidFill = '145a32'
      pt_series.dPt.append(pt_point)
      es_series = ttvc_chart.series[0]
      es_point = DataPoint(idx = 1)
      es_point.graphicalProperties.solidFill = 'F39C12'
      es_series.dPt.append(es_point) 
      #
      #ttvc_ws.add_chart(ttvc_chart, 'D2') 
      #
      #
      # Create charts sheet
      try:
        charts_ws = wb['Result Charts']
      except:
        charts_ws = wb.create_sheet(title = 'Result Charts', index = 0)
      charts_ws.add_chart(mp_chart, 'B2')
      charts_ws.add_chart(ap_chart, 'B24')
      charts_ws.add_chart(tv_chart, 'N2')
      charts_ws.add_chart(ttvc_chart, 'N24')
      #
      #
      # Save file
      wb.save(file_name)
    except Exception as ex:
      print('An error ocurrd while trying to read', file_name)
      print('ERROR ::', ex)
  else:
    print(file_name, 'not found. Unable to generate xlsx charts.')

# ----------- MAIN ----------- #

if __name__ == "__main__":
  # 
  # Get configuration
  config = read_json(sys.argv[1])
  # DEBUG: print(config)
  #
  # Get session results directory
  session_input_dir = config['controller']['temp']['session']['dirs'][ config['controller']['temp']['session']['exec'] ]
  session_results_dir = session_input_dir.replace(config['persistence']['input_base_dir'], config['persistence']['output_base_dir'])
  # DEBUG: print(session_results_dir)
  #
  # Get session results file list
  output_results_file_list = natsort.natsorted(glob.glob(os.path.join(session_results_dir, '*.json')))
  # DEBUG: print(output_results_file_list)
  #
  # For each output file select final results for each period
  final_output_list = []
  for file in output_results_file_list:
    # DEBUG: print(file)
    if '_PT' not in file and '_ES' not in file:
      # Auxiliary
      _PT_ = file.replace('.json', '_PT.json')
      _ES_ = file.replace('.json', '_ES.json')
      #
      # If there's no split add to list
      if not os.path.exists(_PT_) and not os.path.exists(_ES_):
        final_output_list.append(file)
    else:
      # Results after split, add to list
      final_output_list.append(file)
  # DEBUG: print(final_output_list)
  #
  # Initialize lists
  pool_result_list = []
  trading_results_list = []
  # For each final result file
  for file in final_output_list:
    # DEBUG: print(file)
    #
    # Get file content
    content = read_json(file)
    # DEBUG: print(content)
    #
    # Auxiliary variable to get year, month, day, session, period
    elems = file.split(os.sep)[-1].split('.')[0].split('_')
    #
    # Get trading area
    trading_area = 'MI'
    if '_PT' in file:
      trading_area = 'PT'
    elif '_ES' in file:
      trading_area = 'ES'
    #
    # Append data to pool results list
    pool_result_list.append({
      'Year': int(elems[1]),
      'Month': int(elems[2]),
      'Day': int(elems[3]),
      'Session': int(elems[4]),
      'Period': int(elems[5]),
      'Trading Area': trading_area,
      'Pool Result': content['poolResult'],
      'Total Demand (MWh)': round(content['totalDemand'], 1),
      'Total Supply (MWh)': round(content['totalSupply'], 1),
      'Market Price (EUR/MWh)': round(content['marketPrice'], 2),
      'Total Traded Energy (MWh)': round(content['totalTradedEnergy'], 1),
      'Last Demand Trading Unit': content['lastTradedDemandOffer'],
      'Last Supply Trading Unit': content['lastTradedSupplyOffer']
    })
    #
    # For each trading result
    for res in content['tradingResults']:
      # Append data to trading results list
      trading_results_list.append({
        'Year': int(elems[1]),
        'Month': int(elems[2]),
        'Day': int(elems[3]),
        'Session': int(elems[4]),
        'Period': int(elems[5]),
        'Trading Area': trading_area,
        'Unit': res['offerUUID'],
        'Country': 'PT' if res['offerUUID'] in config['controller']['temp']['session']['period']['power_flow']['units']['pt'] else 'ES',
        'Transaction Type': res['transactionType'],
        'Bid Price (EUR/MWh)': round(res['price'], 2),
        'Bid Energy (MWh)': round(res['energy'], 1),
        'Market Price (EUR/MWh)': round(content['marketPrice'], 2),
        'Traded Energy (MWh)': round(res['tradedEnergy'], 1)
      })
  #
  # Set dataframes
  df_pool_results = pd.DataFrame(pool_result_list)
  # DEBUG: print(df_pool_results)
  df_trading_results = pd.DataFrame(trading_results_list)
  # DEBUG: print(df_trading_results)
  #
  # Auxiliary variable to gather chats' data
  market_prices_list = []
  traded_energy_list = []
  # Filter session periods
  period_list = df_pool_results['Period'].unique()
  # DEBUG: print(period_list)
  # For each period result
  for period in period_list:
    # Get period data
    df_temp = df_pool_results[df_pool_results['Period'] == period]
    # DEBUG: print(df_temp)
    if len(df_temp) == 1:
      # MI
      market_prices_list.append({
        'Period': period,
        'PT': df_temp.iloc[0, 9],
        'ES': df_temp.iloc[0, 9],
        'MI': df_temp.iloc[0, 9]
      })
      traded_energy_list.append({
        'Period': period,
        'PT': '',
        'ES': '',
        'MI': df_temp.iloc[0, 10]
      })
    elif len(df_temp) == 2:
      # PT / ES
      market_prices_list.append({
        'Period': period,
        'PT': df_temp[df_temp['Trading Area'] == 'PT'].iloc[0, 9],
        'ES': df_temp[df_temp['Trading Area'] == 'ES'].iloc[0, 9],
        'MI': ''
      })
      traded_energy_list.append({
        'Period': period,
        'PT': df_temp[df_temp['Trading Area'] == 'PT'].iloc[0, 10],
        'ES': df_temp[df_temp['Trading Area'] == 'ES'].iloc[0, 10],
        'MI': ''
      })
  # DEBUG: print(market_prices_list)
  # DEBUG: print(traded_energy_list)
  #
  # Set data frame
  df_market_prices = pd.DataFrame(market_prices_list)
  # DEBUG: print(df_market_prices)
  df_traded_energy = pd.DataFrame(traded_energy_list)
  # DEBUG: print(df_traded_energy)
  #
  # Total Traded Energy By Trading Area 
  total_traded_energy_by_trading_area_list = [{
    'Trading Area': 'Traded Volume PT',
    'Traded Volume (MWh)': round(df_trading_results[df_trading_results['Country'] == 'PT']['Traded Energy (MWh)'].sum(), 1)
  }, {
    'Trading Area': 'Traded Volume ES',
    'Traded Volume (MWh)': round(df_trading_results[df_trading_results['Country'] == 'ES']['Traded Energy (MWh)'].sum(), 1)
  }]
  # DEBUG: print(total_traded_energy_by_trading_area)
  # Set data frame
  df_total_traded_energy_by_trading_area_list = pd.DataFrame(total_traded_energy_by_trading_area_list)
  # DEBUG: print(df_total_traded_energy_by_trading_area_list)
  #
  # Average Prices list
  average_prices_list = [{
    '': 'Submited',
    ' ': 'Mean Price (EUR/MWh)',
    'PT': round(df_trading_results[df_trading_results['Country'] == 'PT']['Bid Price (EUR/MWh)'].mean(), 2),
    'ES': round(df_trading_results[df_trading_results['Country'] == 'ES']['Bid Price (EUR/MWh)'].mean(), 2),
    'MI': round(df_trading_results['Bid Price (EUR/MWh)'].mean(), 2)
  }, {
    '': 'Submited',
    ' ': 'Mean Demand Price (EUR/MWh)',
    'PT': round(df_trading_results[(df_trading_results['Country'] == 'PT') & (df_trading_results['Transaction Type'] == 'buy')]['Bid Price (EUR/MWh)'].mean(), 2),
    'ES': round(df_trading_results[(df_trading_results['Country'] == 'ES') & (df_trading_results['Transaction Type'] == 'buy')]['Bid Price (EUR/MWh)'].mean(), 2),
    'MI': round(df_trading_results[df_trading_results['Transaction Type'] == 'buy']['Bid Price (EUR/MWh)'].mean(), 2)
  }, {
    '': 'Submited',
    ' ': 'Mean Supply Price (EUR/MWh)',
    'PT': round(df_trading_results[(df_trading_results['Country'] == 'PT') & (df_trading_results['Transaction Type'] == 'sell')]['Bid Price (EUR/MWh)'].mean(), 2),
    'ES': round(df_trading_results[(df_trading_results['Country'] == 'ES') & (df_trading_results['Transaction Type'] == 'sell')]['Bid Price (EUR/MWh)'].mean(), 2),
    'MI': round(df_trading_results[df_trading_results['Transaction Type'] == 'sell']['Bid Price (EUR/MWh)'].mean(), 2)
  }, {
    '': 'Clearing',
    ' ': 'Mean Price (EUR/MWh)',
    'PT': round(df_market_prices['PT'].mean(), 2),
    'ES': round(df_market_prices['ES'].mean(), 2),
    'MI': round(df_market_prices[df_market_prices['MI'].values != '']['MI'].sum() / (df_market_prices['MI'].values != '').sum(), 2) if (df_market_prices['MI'].values != '').sum() != 0 else ''
  }]
  # DEBUG: print(average_prices_list)
  # Set data frame
  df_average_prices = pd.DataFrame(average_prices_list)
  # DEBUG: print(df_average_prices)
  # Set excel file name
  sess = session_results_dir.split(os.sep)[-1]
  xlsx_file_name = os.path.join(session_results_dir, 'session_{0}_results.xlsx'.format(sess))
  # DEBUG: print(xlsx_file_name)
  # 
  # Save results to excel
  save_df_to_xlsx_file(
    xlsx_file_name,
    ['Market Prices', 'Mean Prices', 'Traded Volume', 'Total Traded Volume by Country', 'Session Results', 'Trading Results'],
    [df_market_prices, df_average_prices, df_traded_energy, df_total_traded_energy_by_trading_area_list, df_pool_results, df_trading_results]
  )
  # 
  # Draw charts
  draw_charts_xlsx_file(xlsx_file_name)
  #
  # Increment session
  config['controller']['temp']['session']['exec'] += 1
  #
  #
  # rewrite config
  write_json(config, sys.argv[1])